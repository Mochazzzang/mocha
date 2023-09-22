import openpyxl

# 엑셀 파일 불러오기
wb = openpyxl.load_workbook('03. 엑셀 자동화/total.xlsx')

# 현재 활성화 된 시트 선택
ws = wb.active

# 제품이름 리스트
name_list = []

for row in ws.iter_rows(min_row=2, min_col=2):
    name = row[0].value
    if name not in name_list:
        name_list.append(name)
        wb.create_sheet(name)
    data = []
    for cell in row:
        data.append(cell.value)
    wb[name].append(data)        
        
wb.save('03. 엑셀 자동화/total.xlsx')