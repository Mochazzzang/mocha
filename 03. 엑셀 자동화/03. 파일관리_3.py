import openpyxl

save_path = '03. 엑셀 자동화/거래처A 매입 현황.xlsx'

# 기존 엑셀 파일 불러오기
wb = openpyxl.load_workbook(save_path)

# 활성화 된 시트 선택
ws = wb.active

# 데이터 추가 (1)
ws['A1'] = '날짜'
ws['B1'] = '제품명'
ws['C1'] = '가격'
ws['D1'] = '수량'
ws['E1'] = '합계'

# 데이터 추가 (2)
ws.cell(row=2, column=1, value='2030-01-01')
ws.cell(row=2, column=2, value='게이밍 마우스')
ws.cell(row=2, column=3, value=50000)
ws.cell(row=2, column=4, value=30)
ws.cell(row=2, column=5, value='=C2*D2')

# 데이터 추가 (3)
ws.append(['2030-01-03', '기계식 키보드', 120000, 15, '=C3*D3'])

# 데이터 수정
ws['C2'] = 40000
ws['D2'] = 40

# 데이터 삭제
del ws['A3']

# 엑셀 저장
wb.save(save_path)