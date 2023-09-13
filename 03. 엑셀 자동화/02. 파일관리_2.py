import openpyxl

# 새로운 엑셀 파일 생성
wb = openpyxl.Workbook()

# 새로운 시트 생성
ws = wb.create_sheet('2030.01')

# 모든 시트 이름 출력
print(wb.sheetnames)

# 시트 삭제
del wb['Sheet']

# 엑셀 저장
wb.save('03. 엑셀 자동화/거래처A 매입 현황.xlsx')