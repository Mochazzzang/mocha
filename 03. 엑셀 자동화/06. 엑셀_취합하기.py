import openpyxl

# 새로운 엑셀 파일 생성
total_wb = openpyxl.Workbook()

# 현재 활성화된 시트 생성
total_ws = total_wb.active

# 시트 이름 변경
total_ws.title = "data"

# 헤더 추가
total_ws.append(['순번', '제품명', '수량', '합계'])

file_list = ['11번가', '스마트스토어', '쿠팡']

for file in file_list:
    wb = openpyxl.load_workbook(f'03. 엑셀 자동화/{file}.xlsx', data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row = 2):
        data = []
        for cell in row:
            data.append(cell.value)
        total_ws.append(data)
        

total_wb.save('03. 엑셀 자동화/total.xlsx')  