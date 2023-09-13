import openpyxl

save_path = '03. 엑셀 자동화/스마트스토어.xlsx'

# 기존 엑셀 파일 불러오기
wb = openpyxl.load_workbook(save_path, data_only=True)

# data 시트 선택
ws = wb['data']

# 01. 모든 셀 데이터 가져오기
# -> 행과 열 개수를 아는 경우
# for x in range(1, 9+1):
#   for y in range(1, 5+1):
#       print(ws.cell(row=x,column=y).value, end=" ")
#   print()

# -> 행과 열 개수를 모르는 경우
# for x in range(1, ws.max_row+1):
#   for y in range(1, ws.max_column+1):
#       print(ws.cell(row=x,column=y).value, end=" ")
#   print()

# 모든 행 가져오기
# for row in ws.iter_rows():
#     print(row)
    
# 2번째 핼 부터 가져오기
# for row in ws.iter_rows(min_row=2):
#     print(row)
    
# 2번째 행 부터 5번째 행 까지 가져오기
# for row in ws.iter_rows(min_row=2,max_row=5):
#     print(row)
    
# 2-4행, 2-4열 가져오기
for row in ws.iter_rows(min_row=2, max_row=4, min_col=2, max_col=4):
    for cell in row:
      print(cell.value, end=" ")
    print()